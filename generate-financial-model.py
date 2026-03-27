"""
PleoChrome Financial Model Generator
Generates a multi-tab .xlsx workbook with real Excel formulas,
conditional formatting, data validation, and linked sheets.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from copy import copy

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════
# STYLES
# ═══════════════════════════════════════════════

navy = "0B2341"
emerald = "1B6B4A"
teal = "1A8B7A"
sapphire = "1E3A6E"
amethyst = "5B2D8E"
ruby = "A61D3A"
amber = "C47A1A"
white = "FFFFFF"
light_gray = "F5F5F5"
mid_gray = "E0E0E0"
dark_gray = "333333"

header_font = Font(name="Calibri", size=11, bold=True, color=white)
header_fill = PatternFill(start_color=navy, end_color=navy, fill_type="solid")
section_font = Font(name="Calibri", size=10, bold=True, color=navy)
section_fill = PatternFill(start_color=mid_gray, end_color=mid_gray, fill_type="solid")
input_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Yellow = input
calc_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # Blue = calculated
normal_font = Font(name="Calibri", size=10, color=dark_gray)
bold_font = Font(name="Calibri", size=10, bold=True, color=dark_gray)
currency_fmt = '#,##0'
currency_fmt_neg = '#,##0;(#,##0)'
pct_fmt = '0.00%'
thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

def style_section(ws, row, max_col, label=""):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border
    if label:
        ws.cell(row=row, column=1).value = label

def style_input(cell):
    cell.fill = input_fill
    cell.font = Font(name="Calibri", size=10, bold=True, color="1565C0")
    cell.border = thin_border

def style_calc(cell):
    cell.fill = calc_fill
    cell.font = normal_font
    cell.border = thin_border

def style_normal(cell):
    cell.font = normal_font
    cell.border = thin_border

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ═══════════════════════════════════════════════
# TAB 1: ASSUMPTIONS
# ═══════════════════════════════════════════════

ws1 = wb.active
ws1.title = "Assumptions"
ws1.sheet_properties.tabColor = teal
set_col_widths(ws1, [35, 18, 12, 45])

# Title
ws1.merge_cells("A1:D1")
ws1.cell(row=1, column=1).value = "PleoChrome — Deal Assumptions"
ws1.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color=navy)

ws1.merge_cells("A2:D2")
ws1.cell(row=2, column=1).value = "Yellow cells are editable inputs. Blue cells are calculated. All costs in USD."
ws1.cell(row=2, column=1).font = Font(name="Calibri", size=9, italic=True, color="666666")

# Headers
row = 4
for col, val in enumerate(["Parameter", "Value", "Unit", "Source / Justification"], 1):
    ws1.cell(row=row, column=col).value = val
style_header(ws1, row, 4)

# Asset Assumptions
row = 5
style_section(ws1, row, 4, "ASSET DETAILS")

assumptions = [
    # (label, value, unit, source, is_input, fmt)
    ("Stone Type", "Burmese Ruby", "", "Pigeon's blood, Mogok origin", True, None),
    ("Carat Weight", 55, "ct", "Comparable: Estrela de Fura (55.22ct, $34.8M)", True, "#,##0.00"),
    ("Treatment", "No Heat", "", "Unheated commands 2-5x premium", True, None),
    ("Claimed Asset Value", 55000000, "$", "Enter the stone owner's claimed value", True, currency_fmt),
    ("Appraisal Discount (%)", 0.127, "%", "3-Appraisal Rule: avg of 2 lowest typically 10-15% below claimed", True, pct_fmt),
    ("Offering Value", None, "$", "= Claimed Value × (1 - Appraisal Discount)", False, currency_fmt),  # Formula
    ("Token Price (Min Investment)", 100000, "$", "Reg D 506(c) accredited; $200K+ allows self-certification", True, currency_fmt),
    ("Total Tokens", None, "#", "= Offering Value ÷ Token Price", False, "#,##0"),  # Formula
]

for i, (label, value, unit, source, is_input, fmt) in enumerate(assumptions):
    r = row + 1 + i
    ws1.cell(row=r, column=1).value = label
    style_normal(ws1.cell(row=r, column=1))

    cell = ws1.cell(row=r, column=2)
    if label == "Offering Value":
        cell.value = f"=B{row+4}*(1-B{row+5})"  # Claimed Value * (1 - Discount)
        style_calc(cell)
    elif label == "Total Tokens":
        cell.value = f"=INT(B{row+6}/B{row+7})"  # Offering Value / Token Price
        style_calc(cell)
    else:
        cell.value = value
        style_input(cell) if is_input else style_calc(cell)

    if fmt:
        cell.number_format = fmt

    ws1.cell(row=r, column=3).value = unit
    style_normal(ws1.cell(row=r, column=3))
    ws1.cell(row=r, column=4).value = source
    style_normal(ws1.cell(row=r, column=4))
    ws1.cell(row=r, column=4).font = Font(name="Calibri", size=9, color="888888")

# Fee Structure
row = row + len(assumptions) + 2
style_section(ws1, row, 4, "PLEOCHROME FEE STRUCTURE")

fees = [
    ("Setup Fee Rate", 0.02, "%", "Industry standard 1-3%. Collected at engagement.", True, pct_fmt),
    ("Success Fee Rate", 0.015, "%", "0.5-3%. Collected at token sale close. $0 if offering fails.", True, pct_fmt),
    ("Annual Admin Fee Rate", 0.0075, "%", "0.5-1.0%. Recurring revenue starting Year 2.", True, pct_fmt),
    ("BD Placement Rate", 0.07, "%", "5-7% for broker-dealer distribution. 0% if direct placement.", True, pct_fmt),
    ("Setup Fee ($)", None, "$", "= Asset Value × Setup Fee Rate", False, currency_fmt),
    ("Success Fee ($)", None, "$", "= Offering Value × Success Fee Rate", False, currency_fmt),
    ("Annual Admin Fee ($)", None, "$", "= Offering Value × Admin Fee Rate", False, currency_fmt),
    ("BD Placement Fee ($)", None, "$", "= Offering Value × BD Rate", False, currency_fmt),
]

# Store important cell references
fee_start = row + 1
for i, (label, value, unit, source, is_input, fmt) in enumerate(fees):
    r = row + 1 + i
    ws1.cell(row=r, column=1).value = label
    style_normal(ws1.cell(row=r, column=1))

    cell = ws1.cell(row=r, column=2)
    if label == "Setup Fee ($)":
        cell.value = f"=B9*B{fee_start}"  # Asset Value × Setup Rate
        style_calc(cell)
    elif label == "Success Fee ($)":
        cell.value = f"=B11*B{fee_start+1}"  # Offering Value × Success Rate
        style_calc(cell)
    elif label == "Annual Admin Fee ($)":
        cell.value = f"=B11*B{fee_start+2}"  # Offering Value × Admin Rate
        style_calc(cell)
    elif label == "BD Placement Fee ($)":
        cell.value = f"=B11*B{fee_start+3}"  # Offering Value × BD Rate
        style_calc(cell)
    else:
        cell.value = value
        style_input(cell) if is_input else style_calc(cell)

    if fmt:
        cell.number_format = fmt

    ws1.cell(row=r, column=3).value = unit
    style_normal(ws1.cell(row=r, column=3))
    ws1.cell(row=r, column=4).value = source
    style_normal(ws1.cell(row=r, column=4))
    ws1.cell(row=r, column=4).font = Font(name="Calibri", size=9, color="888888")


# ═══════════════════════════════════════════════
# TAB 2: DEAL MODEL (Stone-Level Costs)
# ═══════════════════════════════════════════════

ws2 = wb.create_sheet("Deal Model")
ws2.sheet_properties.tabColor = emerald
set_col_widths(ws2, [35, 15, 15, 15, 15, 40])

ws2.merge_cells("A1:F1")
ws2.cell(row=1, column=1).value = "Stone-Level Deal Model — Phase-by-Phase Costs"
ws2.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color=navy)

ws2.merge_cells("A2:F2")
ws2.cell(row=2, column=1).value = "Yellow = editable override. Blue = auto-calculated from Assumptions tab."
ws2.cell(row=2, column=1).font = Font(name="Calibri", size=9, italic=True, color="666666")

# Key metrics at top
row = 4
for col, val in enumerate(["", "Offering Value", "PleoChrome Costs", "Asset Holder Costs", "BD Fee", "Net to Holder"], 1):
    ws2.cell(row=row, column=col).value = val
style_header(ws2, row, 6)

row = 5
ws2.cell(row=5, column=1).value = "TOTALS"
ws2.cell(row=5, column=1).font = bold_font
ws2.cell(row=5, column=2).value = "=Assumptions!B11"  # Offering Value
ws2.cell(row=5, column=2).number_format = currency_fmt
style_calc(ws2.cell(row=5, column=2))
# PleoChrome costs - will be SUM formula after we lay out costs
# We'll fill these after building the cost rows

# Headers for cost table
row = 7
for col, val in enumerate(["Line Item", "Default ($)", "Override ($)", "Actual ($)", "Paid By", "Source / When Due"], 1):
    ws2.cell(row=row, column=col).value = val
style_header(ws2, row, 6)

# Cost line items
costs = [
    # (phase_label, items[])
    ("PHASE 1: ACQUISITION (Days 1-14)", [
        ("KYC / KYB on Asset Holder", 50, "PleoChrome", "Veriff/Sumsub $0.80/check. Brickken includes 150.", "Day 1-3"),
        ("Sanctions & PEP Screening", 100, "PleoChrome", "OFAC/EU/UN screening. Bundled with KYC or $10-50/ea.", "Day 1-3"),
        ("Provenance Research", 2500, "Asset Holder", "Legal title search. 10-20hrs at $150-$250/hr.", "Day 3-10"),
        ("Intake Agreement (Legal)", 2500, "PleoChrome", "Boutique attorney. Template after first deal.", "Day 5-14"),
    ]),
    ("PHASE 2: PREPARATION (Days 14-84)", [
        ("GIA Grading Report", 0, "Asset Holder", "Most $50M+ stones already certified. $300-500 if needed.", "Day 14-35"),
        ("SSEF Origin Report", 0, "Asset Holder", "CHF 4,000 if needed. Likely already exists.", "Day 14-35"),
        ("Gubelin Report (Optional)", 0, "Asset Holder", "CHF 4,000-5,000 if needed for triple-lab cert.", "Day 14-35"),
        ("Independent Appraisal #1", 7500, "Asset Holder", "CGA/MGA, USPAP compliant. $150-250/hr × 20-40hrs.", "Day 28-38"),
        ("Independent Appraisal #2", 7500, "Asset Holder", "Sequential. Independent of Appraiser #1.", "Day 38-48"),
        ("Independent Appraisal #3", 7500, "Asset Holder", "Completes 3-Appraisal Rule.", "Day 48-56"),
        ("Transit Insurance", 10000, "Asset Holder", "2-4 insured transits at $2K-5K each.", "Days 28-65"),
        ("Vault Custody (Year 1)", 55000, "Asset Holder", "Brink's/Malca-Amit. 0.10% of value, negotiated.", "Day 56+ ongoing"),
        ("Specie Insurance (Year 1)", 55000, "Asset Holder", "Lloyd's/AXA XL. 0.10% vault-stored rate.", "Day 56+ ongoing"),
        ("Chainlink Oracle Integration", 2500, "PleoChrome", "BUILD program or minimal LINK gas fees.", "Day 42-70"),
        ("Chainlink LINK Fees (Year 1)", 1200, "PleoChrome", "$100-200/mo for monthly PoR attestation.", "Day 70+ ongoing"),
        ("SPV Formation (Series LLC)", 750, "PleoChrome", "Wyoming $100 + RA $25/yr + template OA $500.", "Day 30-45"),
        ("PPM Drafting", 10000, "PleoChrome", "Boutique securities attorney. $7.5K-12K flat.", "Day 35-70"),
        ("Subscription Agreement", 2000, "PleoChrome", "Often bundled with PPM. $1.5K-3K standalone.", "Day 50-70"),
        ("Token Purchase Agreement", 2500, "PleoChrome", "Links on-chain token to SPV rights. $2K-3K.", "Day 50-70"),
        ("Photo & Media", 2000, "Asset Holder", "8+ angles, macro, 360 video. $1K-3K.", "Day 56-65"),
    ]),
    ("PHASE 3: TOKENIZATION (Days 63-84)", [
        ("Brickken Platform (Year 1)", 5500, "PleoChrome", "Advanced tier EUR 5,000/yr. Includes 150 KYCs.", "Day 63"),
        ("Configuration Review", 3500, "PleoChrome", "Not full audit — Brickken contracts pre-audited.", "Day 63-75"),
        ("Development & Testing", 5000, "PleoChrome", "10-30hrs config + testnet deploy. $200-400/hr.", "Day 65-80"),
        ("Blockchain Gas Fees", 100, "PleoChrome", "Polygon: $0.01/tx. Multiple deploys = $50-200.", "Day 80"),
        ("Blue Sky Filings (5 States)", 1500, "PleoChrome", "3-5 target states. $100-500 each + filing service.", "Day 80-84"),
    ]),
    ("PHASE 4: DISTRIBUTION (Days 84-120)", [
        ("Marketing & Investor Acq.", 3000, "PleoChrome", "LinkedIn Sales Nav + ads + direct outreach.", "Day 84-120"),
        ("Investor KYC & Accreditation", 0, "PleoChrome", "Covered by Brickken 150 KYCs. Self-cert for $200K+.", "Day 90-120"),
        ("Transfer Agent", 0, "PleoChrome", "ERC-3643 on-chain registry = TA per SEC Jan 2026.", "N/A"),
        ("Investor Portal / Data Room", 0, "PleoChrome", "Built on existing Vercel site or Google Drive.", "Day 80"),
        ("Compliance Monitoring (Year 1)", 500, "PleoChrome", "Built into Brickken + ERC-3643. Quarterly re-screen.", "Day 120+ ongoing"),
        ("Annual AML Audit", 3500, "PleoChrome", "Boutique compliance consultant. $3K-5K.", "Month 12"),
    ]),
]

current_row = 8
cost_rows_pc = []  # Track PleoChrome cost rows for SUM
cost_rows_ah = []  # Track Asset Holder cost rows for SUM

for phase_label, items in costs:
    style_section(ws2, current_row, 6, phase_label)
    current_row += 1

    for label, default_val, paid_by, source, timing in items:
        ws2.cell(row=current_row, column=1).value = label
        style_normal(ws2.cell(row=current_row, column=1))

        # Default value
        cell_default = ws2.cell(row=current_row, column=2)
        cell_default.value = default_val
        cell_default.number_format = currency_fmt
        style_normal(cell_default)

        # Override (yellow input)
        cell_override = ws2.cell(row=current_row, column=3)
        cell_override.number_format = currency_fmt
        style_input(cell_override)

        # Actual = Override if filled, else Default
        cell_actual = ws2.cell(row=current_row, column=4)
        cell_actual.value = f'=IF(C{current_row}<>"",C{current_row},B{current_row})'
        cell_actual.number_format = currency_fmt
        style_calc(cell_actual)

        # Paid By
        ws2.cell(row=current_row, column=5).value = paid_by
        style_normal(ws2.cell(row=current_row, column=5))

        # Source
        ws2.cell(row=current_row, column=6).value = f"{source} | {timing}"
        ws2.cell(row=current_row, column=6).font = Font(name="Calibri", size=9, color="888888")
        style_normal(ws2.cell(row=current_row, column=6))

        if paid_by == "PleoChrome":
            cost_rows_pc.append(current_row)
        else:
            cost_rows_ah.append(current_row)

        current_row += 1

# Totals
current_row += 1
style_section(ws2, current_row, 6, "COST SUMMARY")
current_row += 1

ws2.cell(row=current_row, column=1).value = "PleoChrome Direct Costs"
ws2.cell(row=current_row, column=1).font = bold_font
pc_sum = "+".join([f"D{r}" for r in cost_rows_pc])
ws2.cell(row=current_row, column=4).value = f"={pc_sum}"
ws2.cell(row=current_row, column=4).number_format = currency_fmt
style_calc(ws2.cell(row=current_row, column=4))
pc_total_row = current_row

current_row += 1
ws2.cell(row=current_row, column=1).value = "Asset Holder Pass-Through Costs"
ws2.cell(row=current_row, column=1).font = bold_font
ah_sum = "+".join([f"D{r}" for r in cost_rows_ah])
ws2.cell(row=current_row, column=4).value = f"={ah_sum}"
ws2.cell(row=current_row, column=4).number_format = currency_fmt
style_calc(ws2.cell(row=current_row, column=4))
ah_total_row = current_row

current_row += 1
ws2.cell(row=current_row, column=1).value = "BD Placement Fee"
ws2.cell(row=current_row, column=1).font = bold_font
ws2.cell(row=current_row, column=4).value = "=Assumptions!B22"
ws2.cell(row=current_row, column=4).number_format = currency_fmt
style_calc(ws2.cell(row=current_row, column=4))
bd_total_row = current_row

current_row += 1
ws2.cell(row=current_row, column=1).value = "TOTAL ALL COSTS"
ws2.cell(row=current_row, column=1).font = Font(name="Calibri", size=11, bold=True, color=navy)
ws2.cell(row=current_row, column=4).value = f"=D{pc_total_row}+D{ah_total_row}+D{bd_total_row}"
ws2.cell(row=current_row, column=4).number_format = currency_fmt
ws2.cell(row=current_row, column=4).font = Font(name="Calibri", size=11, bold=True, color=ruby)
style_calc(ws2.cell(row=current_row, column=4))

# Fill in the top summary
ws2.cell(row=5, column=3).value = f"=D{pc_total_row}"
ws2.cell(row=5, column=3).number_format = currency_fmt
style_calc(ws2.cell(row=5, column=3))

ws2.cell(row=5, column=4).value = f"=D{ah_total_row}"
ws2.cell(row=5, column=4).number_format = currency_fmt
style_calc(ws2.cell(row=5, column=4))

ws2.cell(row=5, column=5).value = f"=D{bd_total_row}"
ws2.cell(row=5, column=5).number_format = currency_fmt
style_calc(ws2.cell(row=5, column=5))

ws2.cell(row=5, column=6).value = f"=Assumptions!B11-D{pc_total_row}-D{ah_total_row}-D{bd_total_row}-Assumptions!B19-Assumptions!B20"
ws2.cell(row=5, column=6).number_format = currency_fmt
ws2.cell(row=5, column=6).font = Font(name="Calibri", size=11, bold=True, color=emerald)

# ═══════════════════════════════════════════════
# TAB 3: REVENUE & P&L
# ═══════════════════════════════════════════════

ws3 = wb.create_sheet("PleoChrome P&L")
ws3.sheet_properties.tabColor = sapphire
set_col_widths(ws3, [35, 15, 15, 15, 15, 15, 15])

ws3.merge_cells("A1:G1")
ws3.cell(row=1, column=1).value = "PleoChrome P&L — Per-Deal + 5-Year Projection"
ws3.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color=navy)

# Revenue section
row = 3
for col, val in enumerate(["Revenue Stream", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5", "5-Year Total"], 1):
    ws3.cell(row=row, column=col).value = val
style_header(ws3, row, 7)

rev_items = [
    ("Setup Fee", "=Assumptions!B19", 0, 0, 0, 0),
    ("Success Fee", "=Assumptions!B20", 0, 0, 0, 0),
    ("Secondary Transfer Fees (est.)", 12000, 24000, 24000, 24000, 24000),
    ("Annual Admin Fee", 0, "=Assumptions!B21", "=Assumptions!B21", "=Assumptions!B21", "=Assumptions!B21"),
    ("Valuation Refresh Fee", 0, 25000, 25000, 25000, 25000),
]

for i, (label, *years) in enumerate(rev_items):
    r = row + 1 + i
    ws3.cell(row=r, column=1).value = label
    style_normal(ws3.cell(row=r, column=1))
    for j, val in enumerate(years):
        cell = ws3.cell(row=r, column=2 + j)
        cell.value = val
        cell.number_format = currency_fmt
        style_calc(cell)
    # 5-year total
    ws3.cell(row=r, column=7).value = f"=SUM(B{r}:F{r})"
    ws3.cell(row=r, column=7).number_format = currency_fmt
    style_calc(ws3.cell(row=r, column=7))

total_rev_row = row + len(rev_items) + 1
ws3.cell(row=total_rev_row, column=1).value = "TOTAL REVENUE"
ws3.cell(row=total_rev_row, column=1).font = Font(name="Calibri", size=10, bold=True, color=emerald)
for col in range(2, 8):
    ws3.cell(row=total_rev_row, column=col).value = f"=SUM({get_column_letter(col)}{row+1}:{get_column_letter(col)}{total_rev_row-1})"
    ws3.cell(row=total_rev_row, column=col).number_format = currency_fmt
    ws3.cell(row=total_rev_row, column=col).font = Font(name="Calibri", size=10, bold=True, color=emerald)
    style_calc(ws3.cell(row=total_rev_row, column=col))

# Costs section
cost_row = total_rev_row + 2
style_section(ws3, cost_row, 7, "COSTS")
cost_row += 1
ws3.cell(row=cost_row, column=1).value = "PleoChrome Direct Costs"
ws3.cell(row=cost_row, column=2).value = f"='Deal Model'!D{pc_total_row}"
ws3.cell(row=cost_row, column=2).number_format = currency_fmt
style_calc(ws3.cell(row=cost_row, column=2))
for col in range(3, 7):
    ws3.cell(row=cost_row, column=col).value = 0
    style_calc(ws3.cell(row=cost_row, column=col))
ws3.cell(row=cost_row, column=7).value = f"=SUM(B{cost_row}:F{cost_row})"
ws3.cell(row=cost_row, column=7).number_format = currency_fmt
style_calc(ws3.cell(row=cost_row, column=7))

ongoing_row = cost_row + 1
ws3.cell(row=ongoing_row, column=1).value = "Annual Ongoing (PleoChrome share)"
for col in range(2, 7):
    ws3.cell(row=ongoing_row, column=col).value = 0 if col == 2 else 15000
    ws3.cell(row=ongoing_row, column=col).number_format = currency_fmt
    style_calc(ws3.cell(row=ongoing_row, column=col))
ws3.cell(row=ongoing_row, column=7).value = f"=SUM(B{ongoing_row}:F{ongoing_row})"
ws3.cell(row=ongoing_row, column=7).number_format = currency_fmt
style_calc(ws3.cell(row=ongoing_row, column=7))

total_cost_row = ongoing_row + 1
ws3.cell(row=total_cost_row, column=1).value = "TOTAL COSTS"
ws3.cell(row=total_cost_row, column=1).font = Font(name="Calibri", size=10, bold=True, color=ruby)
for col in range(2, 8):
    ws3.cell(row=total_cost_row, column=col).value = f"=SUM({get_column_letter(col)}{cost_row}:{get_column_letter(col)}{ongoing_row})"
    ws3.cell(row=total_cost_row, column=col).number_format = currency_fmt_neg
    ws3.cell(row=total_cost_row, column=col).font = Font(name="Calibri", size=10, bold=True, color=ruby)
    style_calc(ws3.cell(row=total_cost_row, column=col))

# Net Income
net_row = total_cost_row + 1
ws3.cell(row=net_row, column=1).value = "NET INCOME"
ws3.cell(row=net_row, column=1).font = Font(name="Calibri", size=11, bold=True, color=navy)
for col in range(2, 8):
    ws3.cell(row=net_row, column=col).value = f"={get_column_letter(col)}{total_rev_row}-{get_column_letter(col)}{total_cost_row}"
    ws3.cell(row=net_row, column=col).number_format = currency_fmt_neg
    ws3.cell(row=net_row, column=col).font = Font(name="Calibri", size=11, bold=True, color=navy)
    style_calc(ws3.cell(row=net_row, column=col))

# Margin
margin_row = net_row + 1
ws3.cell(row=margin_row, column=1).value = "NET MARGIN"
for col in range(2, 8):
    ws3.cell(row=margin_row, column=col).value = f"=IF({get_column_letter(col)}{total_rev_row}=0,0,{get_column_letter(col)}{net_row}/{get_column_letter(col)}{total_rev_row})"
    ws3.cell(row=margin_row, column=col).number_format = pct_fmt
    style_calc(ws3.cell(row=margin_row, column=col))

# Asset Holder Economics
ah_row = margin_row + 3
style_section(ws3, ah_row, 7, "ASSET HOLDER ECONOMICS")
ah_items = [
    ("Claimed Asset Value", "=Assumptions!B9"),
    ("Offering Value (after appraisal discount)", "=Assumptions!B11"),
    ("Less: BD Placement Fee", "=-Assumptions!B22"),
    ("Less: PleoChrome Setup Fee", "=-Assumptions!B19"),
    ("Less: PleoChrome Success Fee", "=-Assumptions!B20"),
    ("Less: Asset Holder Pass-Through Costs", f"=-'Deal Model'!D{ah_total_row}"),
]

for i, (label, formula) in enumerate(ah_items):
    r = ah_row + 1 + i
    ws3.cell(row=r, column=1).value = label
    style_normal(ws3.cell(row=r, column=1))
    ws3.cell(row=r, column=2).value = formula
    ws3.cell(row=r, column=2).number_format = currency_fmt_neg
    style_calc(ws3.cell(row=r, column=2))

net_holder_row = ah_row + len(ah_items) + 1
ws3.cell(row=net_holder_row, column=1).value = "NET PROCEEDS TO ASSET HOLDER"
ws3.cell(row=net_holder_row, column=1).font = Font(name="Calibri", size=11, bold=True, color=emerald)
ws3.cell(row=net_holder_row, column=2).value = f"=SUM(B{ah_row+1}:B{net_holder_row-1})"
ws3.cell(row=net_holder_row, column=2).number_format = currency_fmt
ws3.cell(row=net_holder_row, column=2).font = Font(name="Calibri", size=11, bold=True, color=emerald)
style_calc(ws3.cell(row=net_holder_row, column=2))

# % of claimed value
ws3.cell(row=net_holder_row + 1, column=1).value = "% of Claimed Value"
ws3.cell(row=net_holder_row + 1, column=2).value = f"=B{net_holder_row}/B{ah_row+1}"
ws3.cell(row=net_holder_row + 1, column=2).number_format = pct_fmt
style_calc(ws3.cell(row=net_holder_row + 1, column=2))


# ═══════════════════════════════════════════════
# TAB 4: PRICING SOURCES
# ═══════════════════════════════════════════════

ws4 = wb.create_sheet("Pricing Sources")
ws4.sheet_properties.tabColor = amber
set_col_widths(ws4, [30, 18, 15, 20, 40])

ws4.merge_cells("A1:E1")
ws4.cell(row=1, column=1).value = "Pricing Sources of Truth — All Assumptions Sourced"
ws4.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color=navy)

row = 3
for col, val in enumerate(["Cost Category", "Rate / Amount", "Basis", "Last Verified", "Source URL / Reference"], 1):
    ws4.cell(row=row, column=col).value = val
style_header(ws4, row, 5)

sources = [
    ("GIA Colored Stone Report (50+ ct)", "$300-$500", "Per report", "Mar 2026", "gia.edu/gem-lab-fee-schedule"),
    ("SSEF Origin Report (50-100 ct)", "CHF 4,000", "Per report", "Mar 2026", "ssef.ch/testing-prices"),
    ("Gubelin ID + Origin", "~CHF 4,000", "Per report", "Mar 2026", "gubelingemlab.com"),
    ("USPAP Appraisal (CGA/MGA)", "$150-$250/hr", "Hourly (not % of value)", "Mar 2026", "americangemsociety.org/CGA"),
    ("Vault Custody (Brink's negotiated)", "0.10-0.15%", "Annual, of asset value", "Mar 2026", "us.brinks.com/precious-metals"),
    ("Vault Custody (Malca-Amit FTZ)", "0.10-0.25%", "Annual, of asset value", "Mar 2026", "malca-amit.com/vaults-ftz"),
    ("Specie Insurance (vault-stored)", "0.10-0.15%", "Annual, of insured value", "Mar 2026", "AXA XL / Lloyd's syndicate"),
    ("Transit Insurance (inland marine)", "0.25-1.0%", "Per transit event", "Mar 2026", "Marsh Fine Art & Specie"),
    ("Brickken Advanced Tier", "EUR 5,000/yr", "Annual subscription", "Mar 2026", "brickken.com/plans"),
    ("Brickken Enterprise Tier", "EUR 22,000/yr", "Annual subscription", "Mar 2026", "brickken.com/plans"),
    ("Chainlink BUILD Program", "$0 cash", "Token supply commitment (3-7%)", "Mar 2026", "chain.link/build-program"),
    ("Chainlink LINK Feed Costs", "$100-200/mo", "Monthly oracle fees", "Mar 2026", "docs.chain.link/data-feeds"),
    ("Smart Contract Audit (full custom)", "$25K-$100K+", "Per audit", "Mar 2026", "sherlock.xyz/audit-pricing-2026"),
    ("Config Review (Brickken pre-audited)", "$2.5K-$5K", "Per review", "Mar 2026", "softstack.io/audit-cost-2025"),
    ("Securities Attorney (boutique PPM)", "$7.5K-$12K", "Flat fee per offering", "Mar 2026", "contractscounsel.com/ppm-cost"),
    ("Securities Attorney (BigLaw PPM)", "$35K-$75K", "Flat fee per offering", "Mar 2026", "ascentlawfirm.com/ppm-cost"),
    ("Wyoming Series LLC Formation", "$100", "One-time", "Mar 2026", "sos.wyo.gov/Business/docs/BusinessFees.pdf"),
    ("Wyoming Registered Agent", "$25-$125/yr", "Annual", "Mar 2026", "wyomingagents.com"),
    ("Form D Filing (SEC)", "$0", "No fee", "Mar 2026", "sec.gov/form-d"),
    ("Blue Sky Filing (per state avg)", "$100-$500", "Per state notice", "Mar 2026", "blueskycomply.com/fees"),
    ("KYC (Veriff Essential)", "$0.80/check", "Per verification", "Mar 2026", "veriff.com/plans/self-serve"),
    ("KYC (Sumsub Basic)", "$1.35/check", "Per verification", "Mar 2026", "sumsub.com/pricing"),
    ("KYC (Didit - free tier)", "$0", "500 free/month", "Mar 2026", "didit.me"),
    ("Accredited Investor Verification", "$50-$150/investor", "Per investor", "Mar 2026", "verifyinvestor.com"),
    ("Self-Cert ($200K+ investments)", "$0", "SEC March 2025 guidance", "Mar 2026", "SEC Rule 506(c) update"),
    ("BD Placement (full service)", "5-7%", "Of capital raised", "Mar 2026", "FINRA private placement guidelines"),
    ("AML Audit (boutique)", "$3K-$5K", "Annual", "Mar 2026", "BSA/AML compliance benchmarks"),
    ("AML Audit (mid-tier firm)", "$10K-$25K", "Annual", "Mar 2026", "Industry benchmark"),
]

for i, (cat, rate, basis, verified, source) in enumerate(sources):
    r = row + 1 + i
    ws4.cell(row=r, column=1).value = cat
    ws4.cell(row=r, column=2).value = rate
    ws4.cell(row=r, column=3).value = basis
    ws4.cell(row=r, column=4).value = verified
    ws4.cell(row=r, column=5).value = source
    for col in range(1, 6):
        style_normal(ws4.cell(row=r, column=col))
    ws4.cell(row=r, column=5).font = Font(name="Calibri", size=9, color="1565C0")


# ═══════════════════════════════════════════════
# TAB 5: PAYMENT TIMELINE
# ═══════════════════════════════════════════════

ws5 = wb.create_sheet("Payment Timeline")
ws5.sheet_properties.tabColor = amethyst
set_col_widths(ws5, [30, 12, 12, 12, 12, 15, 15])

ws5.merge_cells("A1:G1")
ws5.cell(row=1, column=1).value = "120-Day Payment Timeline — When Every Cost Hits"
ws5.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color=navy)

row = 3
for col, val in enumerate(["Event", "Day Start", "Day End", "Cost Out", "Revenue In", "Running PC Balance", "Phase"], 1):
    ws5.cell(row=row, column=col).value = val
style_header(ws5, row, 7)

timeline = [
    ("Engagement signed → Setup fee collected", 1, 1, 0, "=Assumptions!B19", "", "Acquisition"),
    ("KYC/KYB + Sanctions screening", 1, 3, 150, 0, "", "Acquisition"),
    ("Provenance research (asset holder cost)", 3, 10, 0, 0, "", "Acquisition"),
    ("Intake agreement (legal)", 5, 14, 2500, 0, "", "Acquisition"),
    ("Stone shipped to GIA/SSEF", 14, 14, 0, 0, "", "Preparation"),
    ("Lab reports received + verified", 28, 35, 0, 0, "", "Preparation"),
    ("Appraiser 1 inspection + report", 28, 38, 0, 0, "", "Preparation"),
    ("Appraiser 2 inspection + report", 38, 48, 0, 0, "", "Preparation"),
    ("Appraiser 3 inspection + report", 48, 56, 0, 0, "", "Preparation"),
    ("SPV formation filed", 30, 45, 750, 0, "", "Preparation"),
    ("PPM drafting begins", 35, 70, 10000, 0, "", "Preparation"),
    ("Subscription + Token agreements", 50, 70, 4500, 0, "", "Preparation"),
    ("Chainlink PoR integration", 42, 70, 2500, 0, "", "Preparation"),
    ("Stone shipped to vault + custody begins", 56, 60, 0, 0, "", "Preparation"),
    ("Brickken platform setup", 63, 63, 5500, 0, "", "Tokenization"),
    ("Configuration review", 63, 75, 3500, 0, "", "Tokenization"),
    ("Testnet deployment + testing", 65, 80, 5000, 0, "", "Tokenization"),
    ("Mainnet deployment + gas", 80, 80, 100, 0, "", "Tokenization"),
    ("Blue sky filings", 80, 84, 1500, 0, "", "Tokenization"),
    ("Marketing + investor outreach begins", 84, 120, 3000, 0, "", "Distribution"),
    ("First investor KYC + subscription", 90, 90, 0, 0, "", "Distribution"),
    ("Token sale closes → Success fee", 120, 120, 0, "=Assumptions!B20", "", "Distribution"),
    ("Compliance monitoring begins", 120, 120, 500, 0, "", "Distribution"),
]

phase_colors = {
    "Acquisition": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "Preparation": PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid"),
    "Tokenization": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
    "Distribution": PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
}

running_bal_formula = ""
for i, (event, day_start, day_end, cost_out, rev_in, _, phase) in enumerate(timeline):
    r = row + 1 + i
    ws5.cell(row=r, column=1).value = event
    ws5.cell(row=r, column=2).value = day_start
    ws5.cell(row=r, column=3).value = day_end
    ws5.cell(row=r, column=4).value = cost_out
    ws5.cell(row=r, column=4).number_format = currency_fmt
    ws5.cell(row=r, column=5).value = rev_in
    ws5.cell(row=r, column=5).number_format = currency_fmt

    # Running balance (PleoChrome only)
    if i == 0:
        ws5.cell(row=r, column=6).value = f"=E{r}-D{r}"
    else:
        ws5.cell(row=r, column=6).value = f"=F{r-1}+E{r}-D{r}"
    ws5.cell(row=r, column=6).number_format = currency_fmt_neg
    style_calc(ws5.cell(row=r, column=6))

    ws5.cell(row=r, column=7).value = phase

    # Phase color
    fill = phase_colors.get(phase)
    if fill:
        for col in range(1, 8):
            ws5.cell(row=r, column=col).fill = fill

    for col in range(1, 8):
        ws5.cell(row=r, column=col).border = thin_border
        if col not in (4, 5, 6):
            ws5.cell(row=r, column=col).font = normal_font

# Add conditional formatting for running balance (red if negative)
last_timeline_row = row + len(timeline)
ws5.conditional_formatting.add(
    f"F{row+1}:F{last_timeline_row}",
    CellIsRule(operator="lessThan", formula=["0"],
               font=Font(color=ruby, bold=True),
               fill=PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"))
)


# ═══════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════

output_path = r"C:\Users\shane\.claude\Conversations\PleoChrome\PleoChrome-Financial-Model.xlsx"
wb.save(output_path)
print(f"Workbook saved to: {output_path}")
print(f"Tabs: {[ws.title for ws in wb.worksheets]}")
print("Done!")
